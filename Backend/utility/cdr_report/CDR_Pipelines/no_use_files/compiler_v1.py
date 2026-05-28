import json
import requests
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Border, Side, Font



# ===================== CONFIG =====================

EXCEL_TEMPLATE = "CDR_template.xlsx"
OUTPUT_EXCEL = "cdr_filled.xlsx"
JSON_PATH = "cdr_payload_v5_updated.json"

TABLE_COLUMNS = [
    "photo_no",
    "item_no",
    "name",
    "manufacturer",
    "type_model",
    "technical_data",
    "marks_of_conf",
]

BLUE_FONT = Font(color="0000FF")  # Excel blue
THIN = Side(style="thin")
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OUTSIDE_BORDERS = {"top": Border(top=THIN),"bottom": Border(bottom=THIN),"left": Border(left=THIN),"right": Border(right=THIN),}


# ===================== HELPERS =====================

def split_cell(cell):
    col = "".join(filter(str.isalpha, cell))
    row = int("".join(filter(str.isdigit, cell)))
    return col, row


def merge_if_needed(ws, start_cell, end_cell):
    if start_cell != end_cell:
        ws.merge_cells(f"{start_cell}:{end_cell}")


def download_image(url):
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    return Image(BytesIO(resp.content))

def clear_sheet_from_row(ws, start_row=3):
    """
    Clears all cell values from start_row downward.
    Used for Sheet 3 reset.
    """
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # Also remove images anchored below start_row
    ws._images = [
        img for img in ws._images
        if img.anchor._from.row + 1 < start_row
    ]
    
def set_uniform_row_height_px(ws, start_row, height_px):
    """
    Set row height from start_row downward.
    height_px is in pixels.
    """
    height_points = height_px * 0.75  # px → points

    for r in range(start_row, ws.max_row + 1):
        ws.row_dimensions[r].height = height_points


def resize_image_keep_ratio(img, target_height_px):
    """
    Resize image to target_height_px while preserving aspect ratio.
    """
    ratio = target_height_px / img.height
    img.height = int(target_height_px)
    img.width = int(img.width * ratio)

    


def unmerge_from_row(ws, start_row=3):
    """
    Unmerge all merged cells that touch rows >= start_row.
    """
    merged_ranges = list(ws.merged_cells.ranges)

    for mrange in merged_ranges:
        try:
            min_col, min_row, max_col, max_row = mrange.bounds

            if max_row >= start_row:
                ws.unmerge_cells(str(mrange))

        except Exception:
            # Defensive: skip broken merge metadata
            continue

def find_notes_start_row(ws, start_row=3):
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        val = ws[f"A{r}"].value
        if isinstance(val, str) and val.strip().startswith("NOTES"):
            return r
    return None

def merge_notes_block(ws, notes_start_row, num_rows=4):
    """
    Re-merge NOTES rows from A to G.
    """
    for r in range(notes_start_row, notes_start_row + num_rows):
        ws.merge_cells(f"A{r}:G{r}")

def apply_full_borders(ws, start_row, end_row, start_col="A", end_col="G"):
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)

    for r in range(start_row, end_row + 1):
        for c in range(start_col_idx, end_col_idx + 1):
            ws.cell(row=r, column=c).border = ALL_BORDERS

def apply_outside_borders(ws, start_row, end_row, start_col="A", end_col="G"):
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)

    for r in range(start_row, end_row + 1):
        for c in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            border = Border()

            if r == start_row:
                border = border + OUTSIDE_BORDERS["top"]
            if r == end_row:
                border = border + OUTSIDE_BORDERS["bottom"]
            if c == start_col_idx:
                border = border + OUTSIDE_BORDERS["left"]
            if c == end_col_idx:
                border = border + OUTSIDE_BORDERS["right"]

            cell.border = border


# ===================== SHEET HANDLERS =====================

def handle_sheet_1_2(ws, items):
    """
    Sheet 1 & 2:
    - Write value to answer_cell
    - Clear cell if value is null
    - Skip items without answer_cell
    """
    for item in items:
        answer_cell = item.get("answer_cell")
        if not answer_cell:
            continue

        value = item.get("value")  # may be None
        ws[answer_cell].value = value

        if item.get("value_merged") and item.get("vm_range"):
            merge_if_needed(ws, answer_cell, item["vm_range"])



def handle_sheet_3(ws, items):
    """
    Sheet 3 – FINAL (pixel-correct)
    - Clear content from A3 downward
    - Set row height = 32px
    - Insert images resized to 224px height
    """

    # ---------- 1. Clear existing content ----------
    clear_sheet_from_row(ws, start_row=3)

    # ---------- 2. Set row heights (32px) ----------
    set_uniform_row_height_px(ws, start_row=3, height_px=32)

    # ---------- 3. Rebuild from JSON ----------
    for item in items:
        question_cell = item.get("question_cell")
        if question_cell:
            ws[question_cell].value = item.get("field")

        photo_path = item.get("photo_path")
        answer_cell = item.get("answer_cell")

        if photo_path and answer_cell:
            img = download_image(photo_path)

            # Resize image to 224px height (keep ratio)
            resize_image_keep_ratio(img, target_height_px=224)

            img.anchor = answer_cell
            ws.add_image(img)




def handle_sheet_4(ws, rows):
    """
    Sheet 4 – FINAL, CORRECT, EXCEL-SAFE
    """

    # ---------- 1. Unmerge everything from A3 downward ----------
    unmerge_from_row(ws, start_row=3)

    # ---------- 2. Find NOTES ----------
    notes_start = find_notes_start_row(ws)
    if notes_start is None:
        raise RuntimeError("NOTES section not found in Sheet 4")

    # ---------- 3. Delete existing table rows ----------
    delete_start = 3
    delete_end = notes_start - 2

    if delete_end >= delete_start:
        ws.delete_rows(delete_start, delete_end - delete_start + 1)

    # ---------- 4. Extract table rows ----------
    table_rows = [
        r for r in rows
        if r.get("row_type") == "table_data"
    ]

    # ---------- 5. Insert new rows ----------
    current_row = 3

    for row in table_rows:
        ws.insert_rows(current_row)

        ws[f"A{current_row}"].value = row.get("photo_no")
        ws[f"B{current_row}"].value = row.get("item_no")
        ws[f"C{current_row}"].value = row.get("name")
        ws[f"D{current_row}"].value = row.get("manufacturer")
        ws[f"E{current_row}"].value = row.get("type_model")
        ws[f"F{current_row}"].value = row.get("technical_data")
        ws[f"G{current_row}"].value = row.get("marks_of_conf")

        for col in range(1, 8):  # A to G
            ws.cell(row=current_row, column=col).font = BLUE_FONT
            
        current_row += 1

    # ---------- 6. Re-locate NOTES ----------
    final_notes_start = find_notes_start_row(ws)

    # ---------- 7. Re-merge NOTES ----------
    if final_notes_start is not None:
        merge_notes_block(ws, final_notes_start, num_rows=4)

    # ---------- 8. Apply borders ----------

    final_notes_start = find_notes_start_row(ws)
    if final_notes_start is None:
        return

    table_start = 3
    table_end = final_notes_start - 1

    # Table: full grid borders
    apply_full_borders(ws, table_start, table_end)

    # Notes: outside borders only (5 rows)
    notes_block_start = final_notes_start + 1
    notes_block_end = final_notes_start + 5

    apply_outside_borders(ws, notes_block_start, notes_block_end)

BLUE_FONT_S6 = Font(name="Arial", size=10, color="0000FF")

def handle_sheet_6(ws, items):
    """
    Sheet 6:
    - Unmerge from row 19 downward
    - Clear column B rows 19–33
    - Write to explicit answer_cell
    - Use prefix if value is null
    - Apply field/value merges
    - Apply blue font
    """

    START_ROW = 19
    END_ROW = 33
    TARGET_COL = "B"

    # ---------- 0. Unmerge safely ----------
    unmerge_from_row(ws, start_row=START_ROW)

    # ---------- 1. Clear column B ----------
    for r in range(START_ROW, END_ROW + 1):
        ws[f"{TARGET_COL}{r}"].value = None

    # ---------- 2. Write values + apply merges ----------
    for item in items:
        answer_cell = item.get("answer_cell")
        question_cell = item.get("question_cell")

        if not answer_cell or not question_cell:
            continue

        col, row = split_cell(answer_cell)

        # Safety: only touch B19–B33
        if col != TARGET_COL or not (START_ROW <= row <= END_ROW):
            continue

        value = item.get("value")
        prefix = item.get("prefix")
        final_value = value if value not in (None, "") else prefix

        if final_value in (None, ""):
            continue

        # Write value
        val_cell = ws[answer_cell]
        val_cell.value = final_value
        val_cell.font = BLUE_FONT_S6

        # ---------- Field merge (column A) ----------
        if item.get("field_merged") and item.get("fm_range"):
            merge_if_needed(ws, question_cell, item["fm_range"])

        # ---------- Value merge (column B) ----------
        if item.get("value_merged") and item.get("vm_range"):
            merge_if_needed(ws, answer_cell, item["vm_range"])


# ===================== MAIN =====================

def fill_excel_from_json():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        payload = json.load(f)

    wb = load_workbook(EXCEL_TEMPLATE)

    for sheet_obj in payload["Sheets"]:
        sheet_no = sheet_obj["sheet_no"]
        ws = wb.worksheets[sheet_no - 1]

        # -------- Sheet 1 & 2 --------
        if sheet_no in (1, 2):
            handle_sheet_1_2(ws, sheet_obj.get("Items", []))

        # -------- Sheet 3 --------
        elif sheet_no == 3:
            handle_sheet_3(ws, sheet_obj.get("Items", []))

        # -------- Sheet 4 --------
        elif sheet_no == 4:
            handle_sheet_4(ws, sheet_obj.get("Rows", []))
        
        # -------- Sheet 6 --------
        elif sheet_no == 6:
            handle_sheet_6(ws, sheet_obj.get("Items", []))

        # -------- Any other sheet --------
        else:
            continue

    wb.save(OUTPUT_EXCEL)
    print(f"✅ Excel generated: {OUTPUT_EXCEL}")



fill_excel_from_json()
